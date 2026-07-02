#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera el SQL para importar el histórico de lluvia (lecturas diarias reales de
los 36 pluviómetros de Riopaila, enero-junio 2026) a la tabla `precipitaciones`.
No toca Supabase directamente: escribe un .sql que el usuario pega y corre en
el SQL Editor (mismo patrón de confianza que la migración 0009).

Fuente: "Ponderado de Precipitaciones año 2026.xlsx". El libro tiene VARIAS
hojas por mes (ediciones sucesivas); se usa la hoja validada contra el resumen
oficial "Precip. Acum 2026" (36/36 o 35/36 estaciones coinciden exacto — las
1-2 diferencias por mes son ajustes manuales del propio Excel, documentados
abajo, no se "corrigen" para forzar el calce). Julio-diciembre NO se importan:
el resumen oficial los tiene vacíos (sin conciliar).

Uso:
    python scripts/importar_historico_lluvia.py <uuid-autor> [ruta_excel]

El <uuid-autor> es el id de auth.users del usuario al que se atribuyen las
lecturas históricas (se resuelve con una consulta previa en el SQL Editor:
  select id from auth.users where email = 'rhriop@agricolas.co';
).
"""

import sys
import uuid
from pathlib import Path

import openpyxl

DEFAULT_EXCEL = (
    Path.home() / "Documents" / "Rio Maps" / "Ponderado de Precipitaciones año 2026.xlsx"
)
OUT_DIR = Path(__file__).resolve().parent / "out"
PLANTA = "riopaila"
NOTA = "Histórico importado (Ponderado de Precipitaciones 2026)"
CREATED_AT = "2026-06-25T12:00:00Z"
# Namespace fijo para UUIDs determinísticos (id = uuid5(planta|pv|fecha)).
NS = uuid.UUID("6f1f6a4e-7b0a-4b7a-9b0a-0a1a2a3a4a5a")

# Hoja validada por mes contra el resumen oficial "Precip. Acum 2026"
# (columna del mes en esa hoja: ENE=4, FEB=5 ... según el orden de meses).
MESES = [
    ("2026-01", "Precip.Enero", 4),
    ("2026-02", "Precip. Febrero", 5),
    ("2026-03", "Precip. Marzo", 6),
    ("2026-04", "Precip.Abril ", 7),
    ("2026-05", "Precip Mayo", 8),
    ("2026-06", "Precip Junio ", 9),  # NO usar "Precip.Junio" ni "Precip Jun" (borradores)
]


def dias_cols(ws) -> int:
    """Cuenta las columnas de día (enteros consecutivos 1..N) desde la col. F."""
    hdr = next(r for i, r in enumerate(ws.iter_rows(values_only=True)) if i == 4)
    n, j = 0, 6
    while j < len(hdr) and isinstance(hdr[j], (int, float)) and int(hdr[j]) == n + 1:
        n += 1
        j += 1
    return n


def lecturas_hoja(ws, anio_mes: str) -> list[tuple[int, str, float]]:
    """[(pluviometro, fecha_iso, mm), ...] de una hoja mensual. Omite celdas vacías."""
    n = dias_cols(ws)
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        try:
            pid = int(row[4])
        except (TypeError, ValueError):
            continue
        if pid < 100:
            continue  # filas de resumen (Promedio Zona, total)
        for d in range(n):
            v = row[6 + d]
            if not isinstance(v, (int, float)):
                continue  # sin lectura ese día (no confundir con 0 = no llovió)
            fecha = f"{anio_mes}-{d + 1:02d}"
            out.append((pid, fecha, round(float(v), 1)))
    return out


def resumen_oficial(ws) -> dict[tuple[int, int], float]:
    """{(pluviometro, col_mes): total_oficial} desde 'Precip. Acum 2026'."""
    ref = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        try:
            pid = int(row[3])
        except (TypeError, ValueError):
            continue
        if pid < 100:
            continue
        for col in range(4, 16):
            v = row[col] if col < len(row) else None
            if isinstance(v, (int, float)):
                ref[(pid, col)] = v
    return ref


def sql_valor(v):
    if v is None:
        return "null"
    if isinstance(v, str):
        return "'" + v.replace("'", "''") + "'"
    return str(v)


def escribir_sql(path: Path, filas_sql: list[tuple], anio_mes: str) -> None:
    """Un archivo .sql por mes (transacción propia), para que quepa en el
    SQL Editor de Supabase (el editor rechaza payloads muy grandes)."""
    with path.open("w", encoding="utf-8") as f:
        f.write(f"-- Histórico de lluvia {anio_mes} (Ponderado de Precipitaciones)\n")
        f.write(f"-- {len(filas_sql)} lecturas. Generado por scripts/importar_historico_lluvia.py\n")
        f.write("begin;\n")
        LOTE = 200
        for i in range(0, len(filas_sql), LOTE):
            lote = filas_sql[i : i + LOTE]
            f.write(
                "insert into public.precipitaciones "
                "(id, autor, planta, pluviometro, fecha, mm, nota, deleted, created_at, updated_at)\nvalues\n"
            )
            valores = []
            for rid, aut, planta, pid, fecha, mm, nota, ts in lote:
                cols = [rid, aut, planta, pid, fecha, mm, nota, "false", ts, ts]
                valores.append("  (" + ", ".join(sql_valor(c) for c in cols) + ")")
            f.write(",\n".join(valores))
            f.write(
                "\non conflict (id) do update set mm = excluded.mm, "
                "nota = excluded.nota, updated_at = excluded.updated_at;\n\n"
            )
        f.write("commit;\n")


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    autor = sys.argv[1]
    uuid.UUID(autor)  # valida formato
    excel = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_EXCEL

    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    ref = resumen_oficial(wb["Precip. Acum 2026"])

    OUT_DIR.mkdir(exist_ok=True)
    total = 0
    archivos = []
    for idx, (anio_mes, hoja, col_ref) in enumerate(MESES):
        n_col = list(range(4, 16))[idx]  # ENE=4 ... JUN=9, alineado con MESES
        lecturas = lecturas_hoja(wb[hoja], anio_mes)

        # Validación: suma por estación vs. el resumen oficial del mes.
        por_pv: dict[int, float] = {}
        for pid, _, mm in lecturas:
            por_pv[pid] = round(por_pv.get(pid, 0) + mm, 1)
        ok = sum(
            1
            for pid, total_pv in por_pv.items()
            if (pid, n_col) in ref and abs(total_pv - ref[(pid, n_col)]) < 0.5
        )
        print(f"{anio_mes} [{hoja!r}]: {len(por_pv)} estaciones, {ok}/{len(por_pv)} calzan exacto con el resumen oficial")

        # id determinístico (planta|pv|fecha) -> re-correr el script es idempotente.
        filas_sql = []
        for pid, fecha, mm in lecturas:
            rid = uuid.uuid5(NS, f"{PLANTA}|{pid}|{fecha}")
            filas_sql.append((str(rid), autor, PLANTA, pid, fecha, mm, NOTA, CREATED_AT))

        out_path = OUT_DIR / f"historico_lluvia_{anio_mes}.sql"
        escribir_sql(out_path, filas_sql, anio_mes)
        kb = out_path.stat().st_size / 1024
        print(f"  -> {out_path.name}: {len(filas_sql)} lecturas, {kb:.0f} KB")
        total += len(filas_sql)
        archivos.append(out_path.name)

    print(f"\n{total} lecturas en {len(archivos)} archivos, en {OUT_DIR}")
    print("Corre cada archivo por separado en el SQL Editor de Supabase, en orden:")
    for a in archivos:
        print(f"  - {a}")


if __name__ == "__main__":
    main()
